from breeze_connect import BreezeConnect
from dotenv import load_dotenv
import os
import seckey
import bzins_con
import urllib
import datetime
import pandas as pd
import openpyxl as xl
from openpyxl import load_workbook


def convert_to_excel(data, filename):
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)

def process_trade_data(filename):
    df = pd.read_excel(filename)
    df['trade_date'] = pd.to_datetime(df['trade_date'])
    df["bstp"] = df["average_cost"] / df["quantity"]
    df["bstp"] = df["bstp"].round(2)
    df["investment"] = (
        (df["action"] == "Buy") * (df["average_cost"] + df["brokerage_amount"] + df["total_taxes"])
    )
    df["investment"] = df["investment"].round(2)
    df["revenue"] = (
        (df["action"] == "Sell") * (df["average_cost"] - df["brokerage_amount"] - df["total_taxes"])
    )
    df["revenue"] = df["revenue"].round(2)
    df["pv_investment"] = (
        (df["investment"] > 0) * (df["quantity"] * df["ltp"] * (1 - 0.0065))
    )
    df["pv_investment"] = df["pv_investment"].round(2)
    df["unrealized_profit"] = (df["pv_investment"] - df["investment"])
    df["margin_percent"] = (
        (df["investment"] > 0) * (df["unrealized_profit"] / df["investment"])
    )
    df["margin_percent"] = df["margin_percent"] * 100
    df["margin_percent"] = df["margin_percent"].round(2)
    df["holding_period"] = df["trade_date"].apply(lambda x: (datetime.datetime.now() - x).days / 365)
    df["holding_period"] = df["holding_period"].round(2)
    # Calculate CAGR as a percentage
    df["CAGR"] = (df["unrealized_profit"] / df["investment"] + 1) ** (1 / df["holding_period"]) - 1
    df["CAGR"] = df["CAGR"] * 100  # Convert to percentage 
    df["CAGR"] = df["CAGR"].round(2)
    df = df.sort_values("trade_date", ascending=True)
    return df
# Sort by trade date in ascending order

def write_main_sheet(writer, df):
    df.to_excel(writer, sheet_name="Transactions", index=False)

def write_summary_sheet(writer, df):
    summary_df = df[["stock_code", "exchange_code", "quantity", "investment", "bstp", "pv_investment", "ltp", "unrealized_profit", "margin_percent", "CAGR"]]
    summary_df.to_excel(writer, sheet_name="KeyTrans", index=False)

def write_individual_sheets(writer, df):
    for stock_code in df["stock_code"].unique():
        df_filtered = df[df["stock_code"] == stock_code]
        df_filtered.to_excel(writer, sheet_name=stock_code, index=False)

def calculate_final_set(df):
    final_df = pd.DataFrame()
    for stock_code, group_df in df.groupby("stock_code"):
        buy_quantity = group_df[group_df["action"] == "Buy"]["quantity"].sum()
        sell_quantity = group_df[group_df["action"] == "Sell"]["quantity"].sum()
        bal_quantity =  buy_quantity - sell_quantity
        revenue= group_df[group_df["action"] == "Sell"]["revenue"].sum()
        investment=group_df[group_df["action"] == "Buy"]["investment"].sum()
        ltp = group_df["ltp"].iloc[0]
        pv_bal= ltp * (bal_quantity * (1 - 0.0065))
        bal_cost =  investment - revenue
        if bal_cost < 0:
            bal_cost = 0
            average_price=0
        else:
            average_price = bal_cost / bal_quantity if bal_quantity > 0 else 0
        summary_row = {
            "stock_code": stock_code,
            "exchange_code": group_df["exchange_code"].iloc[0],
            "quantity": bal_quantity,
            "bstp": average_price,
            "investment": investment,
            "revenue": revenue,
            "accrued_cost": bal_cost,
            "ltp": ltp,
            "pv_balance": pv_bal,
            "unrealized_profit": pv_bal - bal_cost,
            "margin_percent": (pv_bal - bal_cost)/bal_cost if bal_cost > 0 else 0
        }
        final_df = pd.concat([final_df, pd.DataFrame([summary_row])], ignore_index=True)
    final_df = final_df.sort_values("margin_percent", ascending=True)
    return final_df

def write_final_set_summary_sheet(writer, final_df):
    final_df.to_excel(writer, sheet_name="Summary", index=False)

def move_summary_sheet_to_front(fin_xl):
    wb = load_workbook(fin_xl)
    sheet_to_move = 'Summary'
    if sheet_to_move in wb.sheetnames:
        wb._sheets = [wb[sheet_to_move]] + [s for s in wb._sheets if s.title != sheet_to_move]
        wb.save(fin_xl)


ses_tok ='49166619'

current_datetime = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
tmp_xl=f'{current_datetime}_trade_list.xlsx'
fin_xl=f'{current_datetime}_trade_report.xlsx'
his_xl=f'{current_datetime}_historical_data.xlsx'
iso_date_string = datetime.datetime.strptime("28/02/2021","%d/%m/%Y").isoformat()[:10] + 'T05:30:00.000Z'
iso_date_time_string = datetime.datetime.strptime("28/02/2021 23:59:59","%d/%m/%Y %H:%M:%S").isoformat()[:19] + '.000Z'
fn_name='stock inventory'
st_rec=bzins_con.secon(ses_tok,fn_name)
convert_to_excel(st_rec, tmp_xl)
df= process_trade_data(tmp_xl)
df.to_excel(fin_xl, index=False)
    # Create an Excel writer object
writer = pd.ExcelWriter(fin_xl, engine="xlsxwriter")
# Write the main sheet
write_main_sheet(writer, df)
# Write the summary sheet
write_summary_sheet(writer, df)
# Write individual sheets for each stock code
write_individual_sheets(writer, df)
# Save the Excel file
writer.close()
# Calculate the final set
final_df = calculate_final_set(df)
# Create a new Excel writer object
writer = pd.ExcelWriter(fin_xl, engine="openpyxl", mode="a")
# Write the final set summary sheet
write_final_set_summary_sheet(writer, final_df)
# Save the workbook
writer.close()
# Move the summary sheet to the front
move_summary_sheet_to_front(fin_xl)


# fn_name='historical data'
# tdate = datetime.datetime.now()
# fdate = tdate - datetime.timedelta(days=1000)
# expiry_date = datetime.date(2023, 3, 29)
# tdate_iso = tdate.isoformat()[:10] + 'T05:30:00.000Z'
# fdate_iso = fdate.isoformat()[:10] + 'T05:30:00.000Z'
# stock_code='YESBAN'
# exchange_code='NSE'
# st_rec=bzins_con.secon(ses_tok,fn_name,start_date=fdate_iso,end_date=tdate_iso,stock_code=stock_code,exchange_code=exchange_code)
# st_rec.to_excel(his_xl, index=False)
